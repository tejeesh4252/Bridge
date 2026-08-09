# =============================================================================
# ACCOUNT LOADER - Reads BMO_Accounts.xlsx dynamically
# =============================================================================

import os
import sys
import logging
import openpyxl
from config import (
    ACCOUNTS_EXCEL_FILENAME,
    ACCOUNTS_SHEET_NAME,
    ACCOUNTS_COL_NAME,
    ACCOUNTS_COL_NUMBER,
    ACCOUNTS_COL_ACTIVE,
    ACCOUNTS_ACTIVE_VALUE,
    ACCOUNTS_HEADER_ROW,
    ACCOUNTS_DATA_START,
)

logger = logging.getLogger("BridgeBMO.AccountLoader")


def get_excel_path() -> str:
    """
    Locate BMO_Accounts.xlsx relative to .exe or script.
    Works in both dev and production.
    """
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(
            os.path.abspath(sys.argv[0])
        )
    return os.path.join(base_dir, ACCOUNTS_EXCEL_FILENAME)


def load_accounts(excel_path: str = None) -> list:
    """
    Read BMO_Accounts.xlsx and return active accounts.

    Returns:
        [{"account_number": "3676145",
          "account_name":   "BC VENTURES JV LP"}, ...]
    """
    path = excel_path or get_excel_path()

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Account file not found:\n{path}\n\n"
            f"Ensure '{ACCOUNTS_EXCEL_FILENAME}' "
            f"is in the same folder as the application."
        )

    try:
        wb = openpyxl.load_workbook(
            path, read_only=True, data_only=True
        )
    except Exception as e:
        raise ValueError(f"Cannot open Excel file:\n{e}")

    if ACCOUNTS_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{ACCOUNTS_SHEET_NAME}' not found.\n"
            f"Available: {', '.join(wb.sheetnames)}"
        )

    ws = wb[ACCOUNTS_SHEET_NAME]

    # ── Read header row ───────────────────────────────────────────────────
    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in next(
            ws.iter_rows(
                min_row=ACCOUNTS_HEADER_ROW,
                max_row=ACCOUNTS_HEADER_ROW
            )
        )
    ]
    logger.info(f"Headers found: {headers}")

    required = [
        ACCOUNTS_COL_NAME,
        ACCOUNTS_COL_NUMBER,
        ACCOUNTS_COL_ACTIVE
    ]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(
            f"Missing columns: {', '.join(missing)}\n"
            f"Expected: {', '.join(required)}"
        )

    col_name   = headers.index(ACCOUNTS_COL_NAME)
    col_number = headers.index(ACCOUNTS_COL_NUMBER)
    col_active = headers.index(ACCOUNTS_COL_ACTIVE)

    # ── Read data rows ────────────────────────────────────────────────────
    accounts = []
    errors   = []

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=ACCOUNTS_DATA_START,
            values_only=True
        ),
        start=ACCOUNTS_DATA_START
    ):
        if not any(row):
            continue

        raw_name   = row[col_name]
        raw_number = row[col_number]
        raw_active = row[col_active]

        name   = str(raw_name).strip()   if raw_name   else ""
        number = str(raw_number).strip() if raw_number else ""

        if number.endswith(".0"):
            number = number[:-2]

        # Silent skip for blank future rows
        if not name and not number:
            continue

        # Active filter
        active_val = str(raw_active).strip() if raw_active else ""
        if active_val.lower() != ACCOUNTS_ACTIVE_VALUE.lower():
            continue

        if not name:
            errors.append(f"Row {row_idx}: Missing Account Name")
            continue

        if not number:
            errors.append(
                f"Row {row_idx}: Missing Account Number"
            )
            continue

        accounts.append({
            "account_number": number,
            "account_name":   name,
        })

    wb.close()

    if errors:
        for e in errors:
            logger.warning(e)

    if not accounts:
        raise ValueError(
            "No active accounts found.\n"
            f"Check '{ACCOUNTS_COL_ACTIVE}' column "
            f"has value '{ACCOUNTS_ACTIVE_VALUE}'."
        )

    logger.info(f"Loaded {len(accounts)} active accounts")
    return accounts


def validate_excel_format(excel_path: str) -> tuple:
    """
    Quick validation. Returns (is_valid, message).
    """
    try:
        accounts = load_accounts(excel_path)
        return True, f"✅   {len(accounts)} active accounts loaded"
    except FileNotFoundError as e:
        return False, f"❌  {str(e)}"
    except ValueError as e:
        return False, f"❌  {str(e)}"
    except Exception as e:
        return False, f"❌  Unexpected error: {str(e)}"
