# =============================================================================
# RECIPIENT LOADER - Updated with Weekly Report filtering
# =============================================================================

import os
import logging
import openpyxl

logger = logging.getLogger("BridgeBMO.Recipients")


def load_recipients(excel_path: str) -> list:
    """
    Load ALL active daily recipients.
    Returns list of email addresses.
    """
    return _load_by_column(
        excel_path,
        active_col   = 1,    # Column B (index 1)
        filter_col   = None,
        filter_value = None
    )


def load_weekly_recipients(excel_path: str) -> list:
    """
    Load recipients marked for Weekly Report.
    Active = Yes AND Weekly Report = Yes.
    Returns list of email addresses.
    """
    return _load_by_column(
        excel_path,
        active_col   = 1,    # Column B - Active
        filter_col   = 3,    # Column D - Weekly Report
        filter_value = "yes"
    )


def _load_by_column(excel_path: str,
                    active_col:   int,
                    filter_col:   int  = None,
                    filter_value: str  = None) -> list:
    """
    Generic recipient loader with optional column filter.
    """
    recipients = []

    if not os.path.exists(excel_path):
        logger.error(f"Excel not found: {excel_path}")
        return []

    try:
        wb = openpyxl.load_workbook(
            excel_path, read_only=True, data_only=True
        )

        if "Recipients" not in wb.sheetnames:
            logger.warning("No Recipients sheet found")
            wb.close()
            return []

        ws = wb["Recipients"]

        for row in ws.iter_rows(
            min_row=4, values_only=True
        ):
            if not row or not row[0]:
                continue

            email  = str(row[0]).strip()
            active = (
                str(row[active_col]).strip().lower()
                if len(row) > active_col and row[active_col]
                else "no"
            )

            # Must be active
            if active != "yes" or "@" not in email:
                continue

            # Apply additional filter if specified
            if filter_col is not None:
                col_val = (
                    str(row[filter_col]).strip().lower()
                    if len(row) > filter_col
                    and row[filter_col]
                    else "no"
                )
                if col_val != filter_value:
                    continue

            recipients.append(email)
            logger.debug(f"Recipient loaded: {email}")

        wb.close()
        logger.info(
            f"Loaded {len(recipients)} recipients "
            f"(filter_col={filter_col})"
        )

    except Exception as e:
        logger.error(f"Recipient load error: {e}")

    return recipients
